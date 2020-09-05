import openpyxl


class HomePageData:
    test_data = [{"firstname":"Anil", "email":"anil@gmail.com", "gender":"Male"}, {"firstname":"Chinni", "email":"chinni@gmail.com", "gender":"Female"}]
    @staticmethod # when we declare static method we do not need self parameter (only for non-static)
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\ANIL\\selenium\\pythonDemo.xlsx")
        sheet = book.active  # to select the active sheet
        Dict = {}
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name: #'TestCase4':  # printinh when only TestCase4 condition satisfies.
                for j in range(2, sheet.max_column + 1):  # to get columns
                    # if you don't want to print testcase4 then start with range 2
                    # print(sheet.cell(row=i, column=j).value)
                    # Dict["lastname"]="Nalla", the below script will build like this
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        # print(Dict)
        return [Dict]
