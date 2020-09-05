from typing import Dict

import openpyxl

book = openpyxl.load_workbook("C:\\Users\\ANIL\\selenium\\pythonDemo.xlsx")
sheet = book.active # to select the active sheet
Dict = {}
cell = sheet.cell(row=1, column=2).value # read
print(cell)

# sheet.cell(row=9, column=9).value = 'Anil' # write back
#
# print(sheet.cell(row=9, column=9).value)

print(sheet.max_row)
print(sheet.max_column)

# print(sheet['A5'].value)

for i in range (1, sheet.max_row+1): # to get rows
    if sheet.cell(row=i, column=1).value == 'TestCase4': # printinh when only TestCase4 condition satisfies.
        for j in range (2, sheet.max_column+1): # to get columns
            # if you don't want to print testcase4 then start with range 2
            # print(sheet.cell(row=i, column=j).value)
            # Dict["lastname"]="Nalla", the below script will build like this
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)